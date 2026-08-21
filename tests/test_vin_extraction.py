"""VIN extraction from spreadsheet cells and free text.

The two failure modes that matter are opposite in direction: inventing a VIN out
of ordinary prose, and silently dropping a real one because of a typo. Both are
tested explicitly.
"""

from __future__ import annotations

import io

import pandas as pd
import pytest

from app.services.vin_extraction import (
    Verdict,
    classify,
    extract_from_dataframe,
    extract_from_text,
    extract_from_workbook,
    to_workbook_bytes,
)

VALID = "5UXKR0C56JL070851"      # 2018 BMW X5, check digit valid
VALID_2 = "WA1ANAFY5J2213924"    # 2018 Audi Q5
VALID_3 = "WBA2J3C53JVA52449"    # 2018 BMW 230i


def reported(text, **kwargs):
    return [c for c in extract_from_text(text, **kwargs) if c.worth_reporting]


class TestLabelledExtraction:
    @pytest.mark.parametrize("cell", [
        f"VIN: {VALID}",
        f"VIN:{VALID}",
        f"VIN {VALID}",
        f"VIN# {VALID}",
        f"VIN#{VALID}",
        f"vin: {VALID}",
        f"CH: {VALID}",
        f"CH# {VALID}",
        f"CH {VALID}",
        f"Chassis: {VALID}",
        f"Chassis No. {VALID}",
        f"CHASSIS NO: {VALID}",
        f"Chassis Number {VALID}",
    ])
    def test_label_forms(self, cell):
        found = reported(cell)
        assert [c.vin for c in found] == [VALID]
        assert found[0].verdict is Verdict.CONFIRMED

    def test_label_is_normalised(self):
        assert reported(f"VIN: {VALID}")[0].label == "VIN"
        assert reported(f"CH: {VALID}")[0].label == "CH"
        assert reported(f"Chassis No. {VALID}")[0].label == "CHASSIS"

    @pytest.mark.parametrize("cell", [
        "VIN: 5UXKR0C5-6JL0-70851",
        "VIN: 5UXKR0C5 6JL0 70851",
        "CH:  5UX KR0 C56J L070851",
        "VIN: 5-U-X-K-R-0-C-5-6-J-L-0-7-0-8-5-1",
    ])
    def test_separators_inside_the_vin(self, cell):
        found = reported(cell)
        assert [c.vin for c in found] == [VALID]

    def test_multiple_vins_in_one_cell(self):
        found = reported(f"VIN: {VALID} and CH: {VALID_2}")
        assert [c.vin for c in found] == [VALID, VALID_2]
        assert [c.label for c in found] == ["VIN", "CH"]

    def test_duplicate_within_a_cell_is_collapsed(self):
        found = reported(f"VIN: {VALID} / VIN: {VALID}")
        assert [c.vin for c in found] == [VALID]

    def test_surrounding_prose_does_not_interfere(self):
        cell = f"Delivered 12/03. VIN: {VALID}. Awaiting inspection at port."
        assert [c.vin for c in reported(cell)] == [VALID]


class TestUnlabelledExtraction:
    def test_bare_vin_is_found(self):
        """A dedicated VIN column has no 'VIN:' prefix in the cell."""
        found = reported(VALID)
        assert [c.vin for c in found] == [VALID]
        assert found[0].label is None

    def test_require_label_skips_bare_vins(self):
        assert extract_from_text(VALID, require_label=True) == []
        assert [c.vin for c in reported(f"VIN: {VALID}", require_label=True)] == [VALID]

    def test_labelled_and_bare_are_not_double_counted(self):
        found = reported(f"VIN: {VALID}")
        assert len(found) == 1


class TestProseIsNotMistakenForVins:
    """The failure mode an unbounded separator regex produces."""

    @pytest.mark.parametrize("sentence", [
        "Delivered to customer, no vin recorded",
        "Vehicle arrived in good condition overall",
        "Awaiting customs clearance at the port terminal",
        "Please contact the sales manager regarding this unit",
        "THE RED CAR WAS PARKED NEAR THE GATE",
    ])
    def test_sentences_produce_nothing_reportable(self, sentence):
        assert reported(sentence) == []

    def test_label_without_a_vin_is_reported_not_silently_dropped(self):
        """A 'CH:' with nothing usable after it means a truncated entry.

        Skipping it quietly would lose a vehicle without telling anyone, so it
        is surfaced for review with an empty VIN and a stated reason.
        """
        found = reported("CH: THE RED CAR WAS PARKED NEAR")
        assert len(found) == 1
        assert found[0].verdict is Verdict.REJECTED
        assert found[0].vin == ""
        assert found[0].label == "CH"
        assert "truncated" in (found[0].reason or "")

    def test_truncated_vin_after_a_label_is_reported(self):
        found = reported("CH: 12345")
        assert len(found) == 1
        assert found[0].vin == ""
        assert "no 17-character VIN" in (found[0].reason or "")

    @pytest.mark.parametrize("sentence", [
        "duplicate of INV-001, hyphenated entry",
        "returned to supplier on 12/03 ref 45",
        "see attached invoice 2024 for details",
    ])
    def test_word_like_runs_with_a_few_digits_are_still_prose(self, sentence):
        """English is full of I and O, so prose reaches the I/O/Q branch too.

        "duplicate of INV-001" collapses to DUPLICATEOFINV001 - 14 letters and
        3 digits. A real VIN is digit-rich; a sentence is letter-rich, and that
        ratio is what keeps the excluded report readable.
        """
        assert reported(sentence) == []

    def test_digit_rich_typo_is_still_surfaced(self):
        """The ratio rule must not hide an actual mistyped VIN."""
        found = reported("1HGCM82633AO04352")
        assert len(found) == 1
        assert found[0].verdict is Verdict.REJECTED

    def test_seventeen_letters_alone_is_not_a_vin(self):
        candidate = classify("ABCDEFGHJKLMNPRST")
        assert candidate.verdict is Verdict.REJECTED


class TestTyposAreSurfaced:
    def test_letter_o_instead_of_zero(self):
        found = reported("VIN: 1HGCM82633AO04352")
        assert len(found) == 1
        assert found[0].verdict is Verdict.REJECTED
        assert "O" in (found[0].reason or "")
        assert "mistyped" in (found[0].reason or "")

    def test_bare_typo_is_still_surfaced(self):
        assert len(reported("1HGCM82633AO04352")) == 1

    def test_failing_check_digit_is_unverified_not_rejected(self):
        """Non-North-American vehicles legitimately fail the check digit."""
        candidate = classify("5UXKR0C57JL070851")     # digit 6 -> 7
        assert candidate.verdict is Verdict.UNVERIFIED
        assert candidate.check_digit_valid is False
        assert candidate.usable is True
        assert "Check digit" in (candidate.reason or "")

    def test_confirmed_has_no_reason(self):
        candidate = classify(VALID)
        assert candidate.verdict is Verdict.CONFIRMED
        assert candidate.reason is None
        assert candidate.check_digit_valid is True


class TestEmptyAndOddInput:
    @pytest.mark.parametrize("value", [None, "", "   ", "\n", 12345, float("nan")])
    def test_no_crash_on_odd_values(self, value):
        assert extract_from_text(value) == []

    def test_short_token_after_a_label_is_flagged_not_treated_as_a_vin(self):
        found = reported("VIN: ABC123")
        assert all(c.vin == "" for c in found)
        assert all(c.verdict is Verdict.REJECTED for c in found)

    def test_eighteen_characters_does_not_yield_a_vin(self):
        """A longer run must not be silently truncated to a fake 17-char VIN."""
        found = reported(f"VIN: {VALID}9")
        assert all(c.verdict is not Verdict.CONFIRMED for c in found)


class TestDataFrameScanning:
    def test_finds_vins_across_columns(self):
        frame = pd.DataFrame({
            "Ref": ["A1", "A2"],
            "Remarks": [f"VIN: {VALID}", "nothing here"],
            "Notes": [f"CH# {VALID_2}", f"chassis no {VALID_3}"],
        })
        result = extract_from_dataframe(frame, "Stock")
        assert set(result.unique_vins) == {VALID, VALID_2, VALID_3}
        assert result.cells_scanned == 6

    def test_row_numbers_match_excel(self):
        """DataFrame row 0 is Excel row 2, because of the header."""
        frame = pd.DataFrame({"Remarks": ["no vin", f"VIN: {VALID}"]})
        result = extract_from_dataframe(frame, "S")
        assert result.occurrences[0].row == 3

    def test_records_sheet_and_column(self):
        frame = pd.DataFrame({"Remarks": [f"VIN: {VALID}"]})
        occurrence = extract_from_dataframe(frame, "Inventory").occurrences[0]
        assert occurrence.sheet == "Inventory"
        assert occurrence.column == "Remarks"
        assert occurrence.original_text == f"VIN: {VALID}"

    def test_same_vin_in_two_rows_is_two_occurrences_one_unique(self):
        frame = pd.DataFrame({"Remarks": [f"VIN: {VALID}", f"CH: {VALID}"]})
        result = extract_from_dataframe(frame, "S")
        assert len(result.occurrences) == 2
        assert result.unique_vins == [VALID]


class TestWorkbookScanning:
    @pytest.fixture
    def workbook(self):
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            pd.DataFrame({
                "Remarks": [f"VIN: {VALID}", "Delivered, no vin recorded"],
            }).to_excel(writer, sheet_name="Sheet1", index=False)
            pd.DataFrame({
                "Chassis": [VALID_2, f"CH# {VALID_3}"],
            }).to_excel(writer, sheet_name="Second Tab", index=False)
        buffer.seek(0)
        buffer.name = "stock.xlsx"
        return buffer

    def test_every_sheet_is_scanned(self, workbook):
        """VINs hide in the second tab more often than anyone would like."""
        result = extract_from_workbook(workbook)
        assert set(result.unique_vins) == {VALID, VALID_2, VALID_3}
        assert "Second Tab" in result.sheets_scanned

    def test_export_has_three_sheets(self, workbook):
        from openpyxl import load_workbook

        result = extract_from_workbook(workbook)
        exported = load_workbook(io.BytesIO(to_workbook_bytes(result)))
        assert exported.sheetnames == ["All VINs", "Unique VINs", "Excluded"]
        assert exported["All VINs"].max_row == len(result.occurrences) + 1

    def test_unreadable_file_reports_an_error_rather_than_raising(self):
        broken = io.BytesIO(b"this is not a spreadsheet")
        broken.name = "broken.xlsx"
        result = extract_from_workbook(broken)
        assert result.errors
        assert result.occurrences == []

    def test_csv_is_supported(self):
        buffer = io.BytesIO(f"ref,remarks\nA1,VIN: {VALID}\n".encode())
        buffer.name = "stock.csv"
        result = extract_from_workbook(buffer)
        assert result.unique_vins == [VALID]


class TestExportShape:
    def test_empty_result_still_produces_a_valid_workbook(self):
        from openpyxl import load_workbook

        from app.services.vin_extraction import ExtractionResult

        exported = load_workbook(io.BytesIO(to_workbook_bytes(ExtractionResult())))
        assert exported.sheetnames == ["All VINs", "Unique VINs", "Excluded"]

    def test_columns_are_not_duplicated(self):
        """The original script listed "VIN" twice in its empty-frame columns."""
        from app.services.vin_extraction import ExtractionResult

        import openpyxl
        exported = openpyxl.load_workbook(io.BytesIO(to_workbook_bytes(ExtractionResult())))
        headers = [c.value for c in exported["All VINs"][1]]
        assert len(headers) == len(set(headers)), f"duplicate column in {headers}"
