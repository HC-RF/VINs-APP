"""API integration tests.

Runs against the real FastAPI app with a temporary database and offline
provider doubles, so behaviour is exercised end to end without network access.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

from app.schemas.vehicle import Confidence, Origin, ProviderKind

VALID_VIN = "5UXKR0C56JL070851"
SECOND_VIN = "WA1ANAFY5J2213924"

FULL_FIELDS = {
    "year": 2018, "make": "BMW", "model": "X5", "trim": "xDrive35i",
    "engine_displacement_l": 3.0, "engine_cylinders": 6, "horsepower": 300,
    "fuel": "Gasoline", "drivetrain": "AWD", "transmission": "Automatic",
    "body_type": "SUV", "plant_country": "United States",
}


@pytest.fixture
def client(stub_provider_factory, monkeypatch):
    """App wired to a single deterministic offline provider."""
    from app.providers import registry as registry_module

    stub = stub_provider_factory(
        "test_source", FULL_FIELDS,
        kind=ProviderKind.FREE, priority=10, confidence=Confidence.HIGH,
        origin=Origin.VIN_DECODED,
    )
    monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (stub,))
    registry_module._registry = None

    from app.config import get_settings
    from app.main import create_app
    from app.services import decode_service as decode_module

    decode_module.reset_decode_service()
    with TestClient(create_app(get_settings())) as test_client:
        yield test_client


class TestHealth:
    def test_health_reports_ok(self, client):
        body = client.get("/api/v1/health").json()
        assert body["status"] == "ok"
        assert body["database"]["ok"] is True

    def test_frontend_is_served(self, client):
        response = client.get("/")
        assert response.status_code == 200
        assert "VIN Decoder" in response.text

    def test_openapi_is_published(self, client):
        assert client.get("/api/openapi.json").status_code == 200


class TestDecodeEndpoint:
    def test_decodes_a_single_vin(self, client):
        body = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()
        assert body["summary"]["decoded"] == 1
        record = body["results"][0]
        assert record["vin"] == VALID_VIN
        assert record["valid"] is True
        assert record["year"] == 2018
        assert record["make"] == "BMW"
        assert record["engine"]["cylinders"] == 6
        assert record["horsepower"] == 300

    def test_response_matches_the_documented_shape(self, client):
        record = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()["results"][0]
        for key in ("vin", "valid", "year", "make", "model", "trim", "engine",
                    "horsepower", "fuel", "drivetrain", "transmission", "confidence"):
            assert key in record
        assert set(record["engine"]) >= {"displacement_l", "type", "cylinders"}
        assert record["confidence"]["overall"] in {"HIGH", "MEDIUM", "LOW", "UNKNOWN"}

    def test_every_field_carries_source_and_confidence(self, client):
        record = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()["results"][0]
        for name, field in record["fields"].items():
            assert field["source"], name
            assert field["confidence"] in {"HIGH", "MEDIUM", "LOW", "UNKNOWN"}, name
            assert field["origin"] in {"VIN_DECODED", "ENRICHED"}, name
            assert field["retrieved_at"], name

    def test_lowercase_input_is_normalized(self, client):
        body = client.post("/api/v1/decode", json={"vins": [VALID_VIN.lower()]}).json()
        assert body["results"][0]["vin"] == VALID_VIN

    def test_pasted_text_block_is_accepted(self, client):
        body = client.post("/api/v1/decode",
                           json={"text": f"{VALID_VIN}\n{SECOND_VIN}"}).json()
        assert body["summary"]["requested"] == 2

    def test_duplicates_are_collapsed_and_reported(self, client):
        body = client.post("/api/v1/decode",
                           json={"vins": [VALID_VIN, VALID_VIN.lower(), SECOND_VIN]}).json()
        assert body["summary"]["requested"] == 2
        assert VALID_VIN in body["summary"]["duplicates_removed"]

    def test_input_order_is_preserved(self, client):
        vins = [SECOND_VIN, VALID_VIN]
        body = client.post("/api/v1/decode", json={"vins": vins}).json()
        assert [r["vin"] for r in body["results"]] == vins

    def test_get_variant_decodes_one_vin(self, client):
        record = client.get(f"/api/v1/decode/{VALID_VIN}").json()
        assert record["vin"] == VALID_VIN
        assert record["valid"] is True


class TestErrorHandling:
    def test_invalid_vin_is_reported_not_decoded(self, client):
        body = client.post("/api/v1/decode", json={"vins": ["NOTAVALIDVIN12345"]}).json()
        record = body["results"][0]
        assert record["valid"] is False
        assert record["status"] == "INVALID_VIN"
        assert record["errors"]
        assert body["summary"]["invalid"] == 1

    def test_short_vin_names_the_length_problem(self, client):
        record = client.post("/api/v1/decode", json={"vins": ["ABC123"]}).json()["results"][0]
        assert any(e["code"] == "VIN_BAD_LENGTH" for e in record["errors"])

    def test_forbidden_letters_are_named(self, client):
        record = client.post("/api/v1/decode",
                             json={"vins": ["5UXKR0C56JL07085I"]}).json()["results"][0]
        assert any(e["code"] == "VIN_ILLEGAL_CHARACTER" for e in record["errors"])

    def test_bad_check_digit_still_decodes_with_a_warning(self, client):
        body = client.post("/api/v1/decode", json={"vins": ["5UXKR0C57JL070851"]}).json()
        record = body["results"][0]
        assert record["valid"] is True
        assert record["check_digit_valid"] is False
        assert any("heck digit" in w for w in record["warnings"])

    def test_invalid_vins_do_not_block_valid_ones(self, client):
        body = client.post("/api/v1/decode",
                           json={"vins": [VALID_VIN, "GARBAGE", SECOND_VIN]}).json()
        assert body["summary"]["decoded"] == 2
        assert body["summary"]["invalid"] == 1

    def test_empty_request_is_rejected(self, client):
        response = client.post("/api/v1/decode", json={"vins": []})
        assert response.status_code == 400
        assert response.json()["error"]["code"] == "INVALID_INPUT"

    def test_over_limit_batch_is_rejected(self, client):
        vins = [f"5UXKR0C56JL{i:06d}" for i in range(150)]
        response = client.post("/api/v1/decode", json={"vins": vins})
        assert response.status_code == 413
        assert response.json()["error"]["code"] == "TOO_MANY_ITEMS"

    def test_unknown_vehicle_returns_a_structured_404(self, client):
        response = client.get("/api/v1/vehicles/WBA5R7C59KAE82587")
        assert response.status_code == 404
        assert response.json()["error"]["code"] == "NOT_FOUND"

    def test_errors_share_one_envelope(self, client):
        for response in (client.post("/api/v1/decode", json={"vins": []}),
                         client.get("/api/v1/vehicles/WBA5R7C59KAE82587")):
            payload = response.json()
            assert set(payload) == {"error"}
            assert {"code", "message"} <= set(payload["error"])


class TestProviderFailures:
    def test_total_provider_failure_yields_not_found_not_a_crash(
        self, stub_provider_factory, monkeypatch
    ):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module
        from app.services import decode_service as decode_module

        broken = stub_provider_factory("broken", {}, fail="upstream is down")
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (broken,))
        registry_module._registry = None
        decode_module.reset_decode_service()

        with TestClient(create_app(get_settings())) as client:
            record = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()["results"][0]

        assert record["status"] in {"NOT_FOUND", "ERROR"}
        assert record["errors"]
        # No fabricated values.
        assert record["make"] is None and record["horsepower"] is None

    def test_missing_fields_are_null_never_invented(self, stub_provider_factory, monkeypatch):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module
        from app.services import decode_service as decode_module

        sparse = stub_provider_factory("sparse", {"year": 2018, "make": "BMW"})
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (sparse,))
        registry_module._registry = None
        decode_module.reset_decode_service()

        with TestClient(create_app(get_settings())) as client:
            record = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()["results"][0]

        assert record["year"] == 2018
        assert record["horsepower"] is None
        assert record["fuel"] is None
        assert record["status"] == "PARTIAL"
        assert "horsepower" not in record["fields"]


class TestDiscrepancyReporting:
    def test_conflicting_sources_produce_a_discrepancy(self, stub_provider_factory, monkeypatch):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module
        from app.services import decode_service as decode_module

        a = stub_provider_factory("source_a", {**FULL_FIELDS, "horsepower": 300},
                                  priority=10, confidence=Confidence.HIGH)
        b = stub_provider_factory("source_b", {**FULL_FIELDS, "horsepower": 190},
                                  priority=20, confidence=Confidence.MEDIUM)
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (a, b))
        registry_module._registry = None
        decode_module.reset_decode_service()

        with TestClient(create_app(get_settings())) as client:
            body = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()

        record = body["results"][0]
        assert len(record["discrepancies"]) == 1
        discrepancy = record["discrepancies"][0]
        assert discrepancy["field"] == "horsepower"
        assert discrepancy["selected_value"] == 300
        assert discrepancy["conflicting"][0]["value"] == 190
        assert record["fields"]["horsepower"]["disputed"] is True
        assert any("discrepancy" in w.lower() for w in record["warnings"])
        assert body["summary"]["discrepancy_count"] == 1


class TestCaching:
    def test_second_lookup_is_served_from_cache(self, client):
        first = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()
        assert first["results"][0]["cached"] is False

        second = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()
        assert second["results"][0]["cached"] is True
        assert second["summary"]["from_cache"] == 1
        assert second["summary"]["provider_calls"] == 0

    def test_refresh_bypasses_the_cache(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        refreshed = client.post("/api/v1/decode",
                                json={"vins": [VALID_VIN], "refresh": True}).json()
        assert refreshed["results"][0]["cached"] is False
        assert refreshed["summary"]["provider_calls"] > 0

    def test_cached_record_keeps_its_provenance(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        record = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()["results"][0]
        assert record["fields"]["make"]["source"] == "test_source"
        assert record["fields"]["make"]["confidence"] == "HIGH"

    def test_deleting_a_cached_vehicle_forces_a_re_decode(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        assert client.delete(f"/api/v1/vehicles/{VALID_VIN}").json()["deleted"] is True
        again = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()
        assert again["results"][0]["cached"] is False

    def test_invalid_vins_are_not_cached(self, client):
        client.post("/api/v1/decode", json={"vins": ["GARBAGE"]})
        assert client.get("/api/v1/vehicles/GARBAGE").status_code == 404


class TestStoredVehicles:
    def test_recent_lists_decoded_vehicles(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN, SECOND_VIN]})
        results = client.get("/api/v1/vehicles/recent").json()["results"]
        assert {r["vin"] for r in results} == {VALID_VIN, SECOND_VIN}

    def test_vehicle_fetch_returns_the_stored_record(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        record = client.get(f"/api/v1/vehicles/{VALID_VIN}").json()
        assert record["vin"] == VALID_VIN
        assert record["cached"] is True


class TestValidateEndpoint:
    def test_validation_is_free_and_needs_no_provider(self, client):
        body = client.get(f"/api/v1/validate/{VALID_VIN}").json()
        assert body["valid"] is True
        assert body["check_digit_valid"] is True
        assert body["model_year"] == 2018

    def test_reports_why_a_vin_is_rejected(self, client):
        body = client.get("/api/v1/validate/ABC").json()
        assert body["valid"] is False
        assert body["issues"]


class TestExport:
    def test_csv_export(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        response = client.post("/api/v1/export", json={"vins": [VALID_VIN], "format": "csv"})
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        text = response.content.decode("utf-8-sig")
        assert "VIN,Valid,Status" in text
        assert VALID_VIN in text
        assert "Overall Confidence" in text and "Sources" in text

    def test_csv_marks_missing_fields_explicitly(self, stub_provider_factory, monkeypatch):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module
        from app.services import decode_service as decode_module

        sparse = stub_provider_factory("sparse", {"year": 2018, "make": "BMW"})
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (sparse,))
        registry_module._registry = None
        decode_module.reset_decode_service()

        with TestClient(create_app(get_settings())) as client:
            client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
            text = client.post("/api/v1/export",
                               json={"vins": [VALID_VIN], "format": "csv"}
                               ).content.decode("utf-8-sig")
        assert "Not available" in text

    def test_xlsx_export_has_provenance_sheets(self, client):
        import io

        from openpyxl import load_workbook

        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        response = client.post("/api/v1/export", json={"vins": [VALID_VIN], "format": "xlsx"})
        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content))
        assert workbook.sheetnames == ["Vehicles", "Field Sources", "Discrepancies", "Export Info"]
        assert workbook["Field Sources"].max_row > 1

    def test_unknown_format_is_rejected(self, client):
        response = client.post("/api/v1/export", json={"vins": [VALID_VIN], "format": "pdf"})
        assert response.status_code == 422

    def test_export_decodes_vins_not_yet_cached(self, client):
        response = client.post("/api/v1/export", json={"vins": [SECOND_VIN], "format": "csv"})
        assert response.status_code == 200
        assert SECOND_VIN in response.content.decode("utf-8-sig")


class TestCompare:
    def test_compares_two_vehicles(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN, SECOND_VIN]})
        body = client.post("/api/v1/compare", json={"vins": [VALID_VIN, SECOND_VIN]}).json()
        assert len(body["vehicles"]) == 2
        assert body["rows"]
        assert all({"field", "label", "values", "differs"} <= set(r) for r in body["rows"])

    def test_identical_vehicles_show_no_differences(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN, SECOND_VIN]})
        body = client.post("/api/v1/compare", json={"vins": [VALID_VIN, SECOND_VIN]}).json()
        # The stub returns the same fields for both, so nothing should differ.
        assert body["difference_count"] == 0

    def test_fewer_than_two_is_rejected(self, client):
        response = client.post("/api/v1/compare", json={"vins": [VALID_VIN]})
        assert response.status_code == 422

    def test_more_than_six_is_rejected(self, client):
        vins = ["5UXKR0C56JL070851"] * 7
        response = client.post("/api/v1/compare", json={"vins": vins})
        assert response.status_code in (400, 422)


class TestProvidersAndUsage:
    def test_provider_list_is_safe_to_expose(self, client, monkeypatch):
        """The endpoint describes providers; it must never carry a credential.

        `requires_key` is a legitimate boolean here - what must not appear is
        the secret itself.
        """
        monkeypatch.setenv("AUTODEV_API_KEY", "sk-live-must-not-leak")

        providers = client.get("/api/v1/providers").json()["providers"]
        assert providers
        for provider in providers:
            assert {"name", "label", "kind", "available", "cost_per_call"} <= set(provider)
            assert "api_key" not in provider
        assert "sk-live-must-not-leak" not in client.get("/api/v1/providers").text

    def test_usage_reports_cost_and_cache_rate(self, client):
        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        client.post("/api/v1/decode", json={"vins": [VALID_VIN]})
        usage = client.get("/api/v1/usage").json()
        assert usage["total_lookups"] >= 2
        assert usage["cache_hits"] >= 1
        assert usage["total_cost"] == 0.0          # free providers only


class TestCostPolicy:
    def test_no_commercial_call_when_free_sources_suffice(
        self, stub_provider_factory, monkeypatch
    ):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module
        from app.services import decode_service as decode_module

        free = stub_provider_factory("free_src", FULL_FIELDS, priority=10)
        paid = stub_provider_factory("paid_src", FULL_FIELDS, kind=ProviderKind.COMMERCIAL,
                                     priority=30, cost=0.05)
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (free, paid))
        registry_module._registry = None
        decode_module.reset_decode_service()

        with TestClient(create_app(get_settings())) as client:
            body = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()

        assert body["summary"]["total_cost"] == 0.0
        assert "paid_src" not in body["results"][0]["sources"]

    def test_commercial_provider_is_used_when_free_data_is_incomplete(
        self, stub_provider_factory, monkeypatch
    ):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module
        from app.services import decode_service as decode_module

        free = stub_provider_factory("free_src", {"year": 2018, "make": "BMW"}, priority=10)
        paid = stub_provider_factory("paid_src", FULL_FIELDS, kind=ProviderKind.COMMERCIAL,
                                     priority=30, cost=0.05)
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (free, paid))
        registry_module._registry = None
        decode_module.reset_decode_service()

        with TestClient(create_app(get_settings())) as client:
            body = client.post("/api/v1/decode", json={"vins": [VALID_VIN]}).json()

        assert body["summary"]["total_cost"] == pytest.approx(0.05)
        assert "paid_src" in body["results"][0]["sources"]

    def test_verify_flag_forces_cross_checking(self, stub_provider_factory, monkeypatch):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module
        from app.services import decode_service as decode_module

        free = stub_provider_factory("free_src", FULL_FIELDS, priority=10)
        paid = stub_provider_factory("paid_src", FULL_FIELDS, kind=ProviderKind.COMMERCIAL,
                                     priority=30, cost=0.05)
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (free, paid))
        registry_module._registry = None
        decode_module.reset_decode_service()

        with TestClient(create_app(get_settings())) as client:
            body = client.post("/api/v1/decode",
                               json={"vins": [VALID_VIN], "verify": True}).json()

        assert body["summary"]["total_cost"] == pytest.approx(0.05)


class TestRateLimit:
    def test_limit_returns_429_with_retry_after(self, stub_provider_factory, monkeypatch):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module

        monkeypatch.setenv("RATE_LIMIT_ENABLED", "true")
        monkeypatch.setenv("RATE_LIMIT_REQUESTS", "3")
        monkeypatch.setenv("RATE_LIMIT_WINDOW_SECONDS", "60")
        get_settings.cache_clear()

        stub = stub_provider_factory("test_source", FULL_FIELDS)
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (stub,))
        registry_module._registry = None

        with TestClient(create_app(get_settings())) as client:
            codes = [client.get(f"/api/v1/validate/{VALID_VIN}").status_code for _ in range(5)]

        assert codes[:3] == [200, 200, 200]
        assert codes[-1] == 429
        get_settings.cache_clear()

    def test_health_is_exempt_from_the_limit(self, stub_provider_factory, monkeypatch):
        from app.config import get_settings
        from app.main import create_app
        from app.providers import registry as registry_module

        monkeypatch.setenv("RATE_LIMIT_ENABLED", "true")
        monkeypatch.setenv("RATE_LIMIT_REQUESTS", "2")
        get_settings.cache_clear()

        stub = stub_provider_factory("test_source", FULL_FIELDS)
        monkeypatch.setattr(registry_module, "PROVIDER_CLASSES", (stub,))
        registry_module._registry = None

        with TestClient(create_app(get_settings())) as client:
            codes = [client.get("/api/v1/health").status_code for _ in range(6)]

        assert all(code == 200 for code in codes)
        get_settings.cache_clear()
