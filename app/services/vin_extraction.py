"""Extract VINs from spreadsheet cells and free text.

Real-world vehicle spreadsheets do not have a tidy VIN column. VINs arrive
buried in remarks fields, prefixed with "VIN:", "CH#" or "Chassis No.", split
by hyphens or spaces, and mixed several to a cell.

The extraction runs in two stages, deliberately:

1. **Find candidates permissively** - any 17-character alphanumeric run,
   tolerating separators. This is intentionally wider than the VIN charset.
2. **Classify each candidate, and say why.** A token containing the letter O is
   almost always a mistyped zero; silently skipping it loses a real vehicle and
   tells the operator nothing. Rejected candidates are returned with a reason
   rather than discarded.

That two-stage split is what stops the extractor from either inventing VINs out
of prose or quietly dropping ones that need a human eye.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import Enum
from typing import Any

from app.vin.validate import VIN_ALLOWED, analyze_vin

#: Labels that introduce a VIN in a spreadsheet cell. `CHASSIS` needs its own
#: alternative because `\bCH\b` will not match inside the longer word.
IDENTIFIER_PATTERN = re.compile(
    r"\b(?:VIN|CH|CHASSIS|CHASIS)\b\s*(?:N[O°]\.?|NUM(?:BER)?|#)?\s*[#:.\-]?\s*",
    re.IGNORECASE,
)

#: A 17-character alphanumeric run, allowing up to two separator characters
#: between adjacent characters so "JTM AB3 FV00 D123456" and
#: "JTMAB3-FV00-D123456" both match.
#:
#: Two details carry weight here:
#:
#: * The `{0,2}` bound. With an unbounded `[\s-]*` the pattern happily stitches
#:   17 letters out of a sentence, turning "THE RED CAR WAS PARKED" into a VIN.
#: * The lookaround anchors. Without them an 18-character run matches its first
#:   17 characters, silently inventing a plausible-looking VIN out of a token
#:   that was never one.
CANDIDATE_PATTERN = re.compile(
    r"(?<![A-Z0-9])[A-Z0-9](?:[\s\-]{0,2}[A-Z0-9]){16}(?![A-Z0-9])"
)

#: Letters that never appear in a VIN; they exist as digits instead.
FORBIDDEN_LETTERS = "IOQ"


#: A field label carries punctuation or a "No./Number" token: "VIN:", "CH#",
#: "Chassis No.". The bare word inside a sentence - "no vin recorded" - does not,
#: and must not be treated as a label with a missing value.
_EXPLICIT_LABEL = re.compile(r"[#:.\-]|\bN[O°]\b|\bNUM", re.IGNORECASE)

#: Invisible characters that survive a copy-paste out of a web page or PDF and
#: split a VIN into two unmatched halves.
_ZERO_WIDTH = re.compile(r"[​-‏⁠﻿]")

#: Unicode dashes. Excel autocorrect turns a typed hyphen into an en dash, which
#: `[\s\-]` does not match, so "5UXKR0C5–6JL070851" would find nothing.
_UNICODE_DASHES = re.compile(r"[‐-―−˗֊᠆]")

#: Non-breaking and exotic spaces.
_ODD_SPACES = re.compile(r"[   -   　]")

#: A label pressed straight against the VIN: "VIN5UXKR0C56JL070851". `\bVIN\b`
#: cannot match there - the boundary fails against the following digit - so the
#: label is invisible and the 20-character run yields nothing.
_GLUED_LABEL = re.compile(
    r"(?<![A-Z0-9])(VIN|CHASSIS|CHASIS|CH)(?=[A-Z0-9]{17}(?![A-Z0-9]))"
)


def normalise_text(raw: object) -> str:
    """Fold a raw cell into the form the matchers expect.

    Every transform here exists because it was observed to hide a real VIN:
    invisible characters split runs, Excel autocorrects hyphens into en dashes,
    and labels get pressed against the number with no separator.
    """
    if raw is None:
        return ""
    text = str(raw)
    text = _ZERO_WIDTH.sub("", text)
    text = _UNICODE_DASHES.sub("-", text)
    text = _ODD_SPACES.sub(" ", text)
    text = text.upper()
    return _GLUED_LABEL.sub(r"\1 ", text)


def _is_explicit_label(matched: str) -> bool:
    return bool(_EXPLICIT_LABEL.search(matched))


def _normalise_label(raw: str) -> str | None:
    """Reduce a matched prefix like "Chassis No. " to a clean label."""
    letters = re.sub(r"[^A-Z]", "", raw.upper())
    if not letters:
        return None
    if letters.startswith("CHAS"):
        return "CHASSIS"
    if letters.startswith("VIN"):
        return "VIN"
    if letters.startswith("CH"):
        return "CH"
    return letters[:8]


class Verdict(str, Enum):
    """How much the extractor trusts a candidate."""

    CONFIRMED = "CONFIRMED"      # valid charset and the check digit validates
    UNVERIFIED = "UNVERIFIED"    # structurally fine, check digit does not match
    REJECTED = "REJECTED"        # not a VIN, with a stated reason


@dataclass(slots=True)
class Candidate:
    """One 17-character token found in a cell."""

    vin: str
    raw: str                      # as it appeared, separators intact
    verdict: Verdict
    reason: str | None = None
    label: str | None = None      # "VIN", "CH", "CHASSIS", or None if unlabelled
    check_digit_valid: bool | None = None
    #: Ordinary prose that merely happened to contain 17 letters. Rejected and
    #: not worth an operator's attention, so it is dropped rather than listed -
    #: otherwise every remarks column would fill the excluded report with noise.
    noise: bool = False

    @property
    def usable(self) -> bool:
        return self.verdict is not Verdict.REJECTED

    @property
    def worth_reporting(self) -> bool:
        return self.usable or not self.noise


@dataclass(slots=True)
class Occurrence:
    """A candidate together with where in the workbook it was found."""

    vin: str
    sheet: str
    row: int                      # 1-based Excel row, header included
    column: str
    original_text: str
    label: str | None
    verdict: Verdict
    reason: str | None
    check_digit_valid: bool | None

    @property
    def cell_ref(self) -> str:
        """Excel-style reference, e.g. ``Stock!B7``."""
        return f"{self.sheet}!{self.column}{self.row}"

    def to_row(self) -> dict[str, Any]:
        return {
            "Cell": self.cell_ref,
            "Sheet": self.sheet,
            "Excel Row": self.row,
            "Column": self.column,
            "VIN": self.vin or "(not found)",
            "Label": self.label or "(unlabelled)",
            "Status": self.verdict.value,
            "Check Digit": (
                "OK" if self.check_digit_valid
                else ("FAILED" if self.check_digit_valid is False else "n/a")
            ),
            "Note": self.reason or "",
            "Original Text": self.original_text,
        }


#: Alphanumeric runs close to VIN length. A run of 14-20 characters that was not
#: accepted is the single most useful diagnostic there is: it is almost always
#: either a truncated VIN, a VIN with a stray character, or a reference number
#: that merely looks like one. Surfacing them turns "some VINs are missing" from
#: a hunch into a list.
NEAR_MISS_PATTERN = re.compile(
    r"(?<![A-Z0-9])[A-Z0-9](?:[\s\-]{0,2}[A-Z0-9]){12,20}(?![A-Z0-9])"
)
NEAR_MISS_MIN = 14
NEAR_MISS_MAX = 20


@dataclass(slots=True)
class NearMiss:
    """A VIN-length-ish run that was not accepted, and how long it actually is."""

    token: str
    length: int
    sheet: str
    row: int
    column: str
    original_text: str

    def to_row(self) -> dict[str, Any]:
        return {
            "Cell": f"{self.sheet}!{self.column}{self.row}",
            "Sheet": self.sheet,
            "Token": self.token,
            "Length": self.length,
            "Why not a VIN": (
                f"{self.length} characters, a VIN has 17 "
                f"({'too short' if self.length < 17 else 'too long'})"
            ),
            "Original Text": self.original_text,
        }


@dataclass
class ExtractionResult:
    occurrences: list[Occurrence] = field(default_factory=list)
    rejected: list[Occurrence] = field(default_factory=list)
    near_misses: list[NearMiss] = field(default_factory=list)
    sheets_scanned: list[str] = field(default_factory=list)
    cells_scanned: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def unique_vins(self) -> list[str]:
        """Distinct usable VINs, in first-seen order."""
        seen: dict[str, None] = {}
        for occ in self.occurrences:
            seen.setdefault(occ.vin, None)
        return list(seen)

    @property
    def confirmed_count(self) -> int:
        return sum(1 for o in self.occurrences if o.verdict is Verdict.CONFIRMED)

    @property
    def unverified_count(self) -> int:
        return sum(1 for o in self.occurrences if o.verdict is Verdict.UNVERIFIED)


# --- Candidate classification ----------------------------------------------

def classify(raw: str, *, label: str | None = None) -> Candidate:
    """Turn a raw 17-character run into a classified candidate."""
    vin = re.sub(r"[\s\-]", "", raw).upper()
    digits = sum(1 for c in vin if c.isdigit())

    # Decided up front, because prose is usually rejected by the I/O/Q check
    # long before the digit test would ever run. Every real VIN carries a
    # check digit in position 9, so a 17-character run with no digits at all
    # and no "VIN:"/"CH:" label in front of it is a sentence, not a vehicle.
    prose = label is None and digits == 0

    if len(vin) != 17:
        return Candidate(vin, raw, Verdict.REJECTED,
                         f"{len(vin)} characters after removing separators, not 17.",
                         label, noise=prose)

    bad_letters = sorted({c for c in vin if c in FORBIDDEN_LETTERS})
    if bad_letters:
        # Very often a transcription slip: O for 0, I for 1 - worth surfacing.
        #
        # But unlabelled prose reaches this branch too, because English is full
        # of I and O: "duplicate of INV-001" collapses to DUPLICATEOFINV001.
        # A real VIN is digit-rich (17 characters carrying a serial number);
        # a sentence is letter-rich. That ratio separates them cleanly.
        word_like = label is None and digits <= 3
        return Candidate(
            vin, raw, Verdict.REJECTED,
            f"Contains {', '.join(bad_letters)}, which never appear in a VIN "
            f"(likely a mistyped 0 or 1).",
            label, noise=prose or word_like,
        )

    if any(c not in VIN_ALLOWED for c in vin):
        return Candidate(vin, raw, Verdict.REJECTED,
                         "Contains characters outside the VIN charset.",
                         label, noise=prose)

    analysis = analyze_vin(vin)
    check_ok = analysis.check_digit_valid

    if check_ok:
        return Candidate(vin, raw, Verdict.CONFIRMED, None, label, True)

    # The check digit did not validate. That is normal for vehicles not built to
    # the North American standard, so it is not disqualifying on its own - but a
    # token with no digits at all is prose, not a VIN.
    if digits == 0:
        return Candidate(
            vin, raw, Verdict.REJECTED,
            "No digits and the check digit does not validate; this is almost "
            "certainly ordinary text, not a VIN.",
            label, False,
            # An unlabelled digit-free run is just a sentence. A labelled one
            # means a "VIN:" was found without a usable VIN after it, which the
            # operator does want to see.
            noise=label is None,
        )
    if digits < 2 and label is None:
        return Candidate(
            vin, raw, Verdict.REJECTED,
            "Unlabelled, nearly all letters, and the check digit does not "
            "validate; treated as text rather than a VIN.",
            label, False, noise=True,
        )

    return Candidate(
        vin, raw, Verdict.UNVERIFIED,
        f"Check digit does not validate (expected "
        f"'{analysis.expected_check_digit}', found '{analysis.actual_check_digit}'). "
        f"Common for imported vehicles; verify against the document.",
        label, False,
    )


def extract_from_text(text: Any, *, require_label: bool = False) -> list[Candidate]:
    """Find every VIN candidate in a single cell.

    When ``require_label`` is true, only runs introduced by VIN/CH/CHASSIS are
    considered. Otherwise a bare 17-character token also counts - which is what
    a dedicated VIN column looks like.
    """
    if text is None:
        return []
    raw_text = str(text)
    if not raw_text.strip():
        return []
    upper = normalise_text(raw_text)

    found: list[Candidate] = []
    seen: set[str] = set()
    covered: list[tuple[int, int]] = []

    # --- Pass 1: labelled candidates ---------------------------------------
    matches = list(IDENTIFIER_PATTERN.finditer(upper))
    for i, match in enumerate(matches):
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(upper)
        section = upper[start:end]

        label = _normalise_label(match.group(0))

        # Only the FIRST run after the label belongs to it. "VIN: <vin>. Awaiting
        # inspection at port." must not hand the label to the trailing sentence
        # and report "AWAITINGINSPECTIO" as a VIN. Anything genuinely further
        # along is picked up unlabelled by pass 2, where the prose filter applies.
        candidate_match = CANDIDATE_PATTERN.search(section)
        if candidate_match is None:
            # The cell says "VIN:" or "CH:" but nothing usable follows - a
            # truncated or mistyped entry. Silently skipping it loses a vehicle
            # without telling anyone, so it is surfaced for review instead.
            #
            # Only for an explicit field label though: the bare word in
            # "no vin recorded" is prose, not an empty field.
            if not _is_explicit_label(match.group(0)):
                continue
            found.append(Candidate(
                vin="", raw=section.strip()[:40], verdict=Verdict.REJECTED,
                reason=f"A '{label}' label was found but no 17-character VIN "
                       f"follows it; the entry may be truncated.",
                label=label,
            ))
            continue

        candidate = classify(candidate_match.group(0), label=label)
        covered.append((start + candidate_match.start(), start + candidate_match.end()))
        if candidate.vin in seen:
            continue
        seen.add(candidate.vin)
        found.append(candidate)

    if require_label:
        return found

    # --- Pass 2: bare candidates, skipping regions already consumed --------
    for candidate_match in CANDIDATE_PATTERN.finditer(upper):
        span = candidate_match.span()
        if any(span[0] < c_end and c_start < span[1] for c_start, c_end in covered):
            continue
        candidate = classify(candidate_match.group(0), label=None)
        if candidate.vin in seen:
            continue
        seen.add(candidate.vin)
        found.append(candidate)

    return found


# --- Workbook scanning ------------------------------------------------------

def extract_from_dataframe(
    frame, sheet_name: str = "Sheet1", *, require_label: bool = False,
    header_offset: int = 2,
) -> ExtractionResult:
    """Scan every cell of a DataFrame.

    ``header_offset`` converts a 0-based DataFrame index to the row number the
    user sees in Excel: 2 accounts for the header row plus 1-based numbering.
    """
    result = ExtractionResult(sheets_scanned=[sheet_name])

    # Column names are data too when the file had no header row.
    for col_index, column in enumerate(frame.columns, start=1):
        result.cells_scanned += 1
        for candidate in extract_from_text(column, require_label=require_label):
            occurrence = Occurrence(
                vin=candidate.vin, sheet=sheet_name, row=1,
                column=_column_letter(col_index), original_text=str(column),
                label=candidate.label, verdict=candidate.verdict,
                reason=candidate.reason, check_digit_valid=candidate.check_digit_valid,
            )
            if candidate.usable:
                result.occurrences.append(occurrence)
            elif candidate.worth_reporting:
                result.rejected.append(occurrence)

    for row_index, row in frame.iterrows():
        for column in frame.columns:
            value = row[column]
            result.cells_scanned += 1
            for candidate in extract_from_text(value, require_label=require_label):
                occurrence = Occurrence(
                    vin=candidate.vin,
                    sheet=sheet_name,
                    row=(row_index if isinstance(row_index, int) else 0) + header_offset,
                    column=str(column),
                    original_text=str(value),
                    label=candidate.label,
                    verdict=candidate.verdict,
                    reason=candidate.reason,
                    check_digit_valid=candidate.check_digit_valid,
                )
                if candidate.usable:
                    result.occurrences.append(occurrence)
                elif candidate.worth_reporting:
                    result.rejected.append(occurrence)

    return result


def _column_letter(index: int) -> str:
    """1 -> A, 27 -> AA. Excel's own column naming."""
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def iter_cells(source):
    """Yield ``(sheet, row, column, value)`` for every populated cell.

    Deliberately **not** built on ``pandas.read_excel``. That function treats
    the first row as column names, so any VIN sitting on row 1 - which is every
    file exported without a header - is consumed into the frame's columns and
    never scanned. Reading raw cells means no row is special, and the position
    reported back is a real Excel reference like ``Sheet1!B7``.
    """
    name = str(getattr(source, "name", source) or "").lower()

    if name.endswith((".csv", ".txt", ".tsv")):
        yield from _iter_delimited(source, name)
        return

    from openpyxl import load_workbook

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    yield (
                        worksheet.title,
                        cell.row,
                        _column_letter(cell.column),
                        cell.value,
                    )
    finally:
        workbook.close()


def _iter_delimited(source, name: str):
    """CSV/TSV, read as raw rows with no header row and no type coercion."""
    import csv
    import io as _io

    data = source.read() if hasattr(source, "read") else open(source, "rb").read()
    if isinstance(data, bytes):
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = data.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = data.decode("utf-8", errors="replace")
    else:
        text = data

    delimiter = "\t" if name.endswith(".tsv") else ","
    if not name.endswith(".tsv"):
        try:
            delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","

    sheet = "CSV"
    for row_index, row in enumerate(csv.reader(_io.StringIO(text), delimiter=delimiter), start=1):
        for col_index, value in enumerate(row, start=1):
            if value is None or not str(value).strip():
                continue
            yield sheet, row_index, _column_letter(col_index), value


def extract_from_workbook(source, *, require_label: bool = False) -> ExtractionResult:
    """Scan every cell of every sheet.

    ``source`` is a path or a file-like object - Streamlit's uploaded file works
    directly. Nothing is treated as a header, so row 1 is scanned like any other.
    """
    combined = ExtractionResult()
    seen_sheets: dict[str, None] = {}

    try:
        cells = list(iter_cells(source))
    except Exception as exc:  # noqa: BLE001 - surfaced to the user, not raised
        combined.errors.append(
            f"Could not read the file: {exc}. "
            f"Supported formats are .xlsx, .xlsm, .csv, .tsv and .txt "
            f"(the older .xls format must be re-saved as .xlsx)."
        )
        return combined

    if not cells:
        combined.errors.append("The file contains no readable cells.")
        return combined

    for sheet, row, column, value in cells:
        seen_sheets.setdefault(sheet, None)
        combined.cells_scanned += 1

        candidates = extract_from_text(value, require_label=require_label)
        handled: set[str] = set()

        for candidate in candidates:
            handled.add(candidate.vin)
            occurrence = Occurrence(
                vin=candidate.vin,
                sheet=sheet,
                row=row,
                column=column,
                original_text=str(value),
                label=candidate.label,
                verdict=candidate.verdict,
                reason=candidate.reason,
                check_digit_valid=candidate.check_digit_valid,
            )
            if candidate.usable:
                combined.occurrences.append(occurrence)
            elif candidate.worth_reporting:
                combined.rejected.append(occurrence)

        for token, length in find_near_misses(value, handled):
            combined.near_misses.append(
                NearMiss(token, length, sheet, row, column, str(value))
            )

    combined.sheets_scanned = list(seen_sheets)
    return combined


def find_near_misses(text: Any, handled: set[str]) -> list[tuple[str, int]]:
    """Runs that are nearly VIN-length but were not accepted.

    This is the answer to "are we missing any?". A 16- or 18-character run in a
    VIN column is almost certainly a real vehicle with a dropped or doubled
    character, and it would otherwise vanish without trace.
    """
    if text is None:
        return []
    normalised = normalise_text(text)
    if not normalised.strip():
        return []

    out: list[tuple[str, int]] = []
    seen: set[str] = set()
    for match in NEAR_MISS_PATTERN.finditer(normalised):
        token = re.sub(r"[\s\-]", "", match.group(0))
        if len(token) == 17 or not (NEAR_MISS_MIN <= len(token) <= NEAR_MISS_MAX):
            continue                       # 17 was already classified above
        if token in handled or token in seen:
            continue
        # A run that swallows a VIN we already extracted is not a miss. This is
        # what "VIN5UXKR0C56JL070851" looks like once the label is split off:
        # the 17 characters are reported, and the 20-character span around them
        # would otherwise be listed as though something had been overlooked.
        if any(found and found in token for found in handled):
            continue
        # A run made almost entirely of letters is a word, not a broken VIN.
        if sum(1 for c in token if c.isdigit()) == 0:
            continue
        seen.add(token)
        out.append((token, len(token)))
    return out


# --- Export -----------------------------------------------------------------

def to_workbook_bytes(result: ExtractionResult) -> bytes:
    """Build the two-sheet Excel export: every occurrence, and unique VINs."""
    import io

    import pandas as pd

    columns = ["Cell", "Sheet", "Excel Row", "Column", "VIN", "Label", "Status",
               "Check Digit", "Note", "Original Text"]

    all_rows = [o.to_row() for o in result.occurrences]
    all_frame = pd.DataFrame(all_rows, columns=columns) if all_rows else pd.DataFrame(columns=columns)
    unique_frame = all_frame.drop_duplicates(subset=["VIN"]).copy()

    rejected_rows = [o.to_row() for o in result.rejected]
    rejected_frame = (
        pd.DataFrame(rejected_rows, columns=columns) if rejected_rows
        else pd.DataFrame([{**{c: "" for c in columns},
                            "Note": "No candidates were rejected."}], columns=columns)
    )

    near_columns = ["Cell", "Sheet", "Token", "Length", "Why not a VIN", "Original Text"]
    near_rows = [n.to_row() for n in result.near_misses]
    near_frame = (
        pd.DataFrame(near_rows, columns=near_columns) if near_rows
        else pd.DataFrame([{**{c: "" for c in near_columns},
                            "Why not a VIN": "No near misses - nothing of VIN-like "
                                             "length went unaccounted for."}],
                          columns=near_columns)
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        all_frame.to_excel(writer, sheet_name="All VINs", index=False)
        unique_frame.to_excel(writer, sheet_name="Unique VINs", index=False)
        rejected_frame.to_excel(writer, sheet_name="Excluded", index=False)
        near_frame.to_excel(writer, sheet_name="Near Misses", index=False)
    return buffer.getvalue()
