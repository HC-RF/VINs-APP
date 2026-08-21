"""CSV and Excel export.

Exports carry provenance, not just values. A spreadsheet that says "300 hp"
without saying where the number came from is exactly the artefact this
application exists to avoid, so every export includes a Sources sheet/column
set and flags disputed fields.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.vehicle import VehicleRecord

#: (column header, attribute path) for the flat tabular views.
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("VIN", "vin"),
    ("Valid", "valid"),
    ("Status", "status"),
    ("Year", "year"),
    ("Make", "make"),
    ("Model", "model"),
    ("Trim", "trim"),
    ("Body Type", "body_type"),
    ("Engine (L)", "engine.displacement_l"),
    ("Engine Type", "engine.type"),
    ("Cylinders", "engine.cylinders"),
    ("Horsepower", "horsepower"),
    ("Torque (lb-ft)", "engine.torque_lb_ft"),
    ("Fuel", "fuel"),
    ("Drivetrain", "drivetrain"),
    ("Transmission", "transmission"),
    ("Transmission Speeds", "transmission_speeds"),
    ("MPG City", "mpg_city"),
    ("MPG Highway", "mpg_highway"),
    ("MPG Combined", "mpg_combined"),
    ("Manufacturer", "manufacturer"),
    ("Country of Manufacture", "plant_country"),
    ("Plant City", "plant_city"),
    ("Overall Confidence", "_confidence"),
    ("Sources", "_sources"),
    ("Discrepancies", "_discrepancy_count"),
    ("Discrepancy Detail", "_discrepancy_detail"),
    ("Warnings", "_warnings"),
    ("Cached", "cached"),
    ("Decoded At (UTC)", "_decoded_at"),
]

NOT_AVAILABLE = "Not available"


def _resolve(record: VehicleRecord, path: str):
    """Read a value for an export column, including computed ``_`` columns."""
    if path == "_confidence":
        return record.confidence.get("overall", "UNKNOWN")
    if path == "_sources":
        return ", ".join(record.sources) if record.sources else NOT_AVAILABLE
    if path == "_discrepancy_count":
        return len(record.discrepancies)
    if path == "_discrepancy_detail":
        return " | ".join(d.message for d in record.discrepancies) or ""
    if path == "_warnings":
        return " | ".join(record.warnings) or ""
    if path == "_decoded_at":
        return record.decoded_at.strftime("%Y-%m-%d %H:%M:%S")

    node = record
    for part in path.split("."):
        node = getattr(node, part, None)
        if node is None:
            return NOT_AVAILABLE
    if hasattr(node, "value"):          # enum
        return node.value
    return node


def _cell(value) -> object:
    if value is None:
        return NOT_AVAILABLE
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def to_csv(records: list[VehicleRecord]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([header for header, _ in EXPORT_COLUMNS])
    for record in records:
        writer.writerow([_cell(_resolve(record, path)) for _, path in EXPORT_COLUMNS])
    return buffer.getvalue()


# --- Excel ------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_DISPUTE_FILL = PatternFill("solid", fgColor="FEF3C7")
_INVALID_FILL = PatternFill("solid", fgColor="FEE2E2")
_TITLE_FONT = Font(bold=True, size=13)


def _style_header(ws, columns: list[str]) -> None:
    for idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize(ws, max_width: int = 52) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(10, longest + 2), max_width)


def to_xlsx(records: list[VehicleRecord]) -> bytes:
    wb = Workbook()

    # --- Sheet 1: the vehicles ---------------------------------------------
    ws = wb.active
    ws.title = "Vehicles"
    headers = [header for header, _ in EXPORT_COLUMNS]
    _style_header(ws, headers)

    for record in records:
        row = [_cell(_resolve(record, path)) for _, path in EXPORT_COLUMNS]
        ws.append(row)
        r = ws.max_row
        if not record.valid:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = _INVALID_FILL
        elif record.discrepancies:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = _DISPUTE_FILL
    _autosize(ws)

    # --- Sheet 2: field-level provenance -----------------------------------
    ws2 = wb.create_sheet("Field Sources")
    prov_headers = ["VIN", "Field", "Value", "Source", "Source Type",
                    "Confidence", "Origin", "Disputed", "Retrieved (UTC)", "Note"]
    _style_header(ws2, prov_headers)
    for record in records:
        for name, field in sorted(record.fields.items()):
            if field.value is None:
                continue
            ws2.append([
                record.vin,
                field.label or name,
                _cell(field.value),
                field.source or NOT_AVAILABLE,
                field.source_kind.value if field.source_kind else "",
                field.confidence.value,
                field.origin.value if field.origin else "",
                "Yes" if field.disputed else "No",
                field.retrieved_at.strftime("%Y-%m-%d %H:%M:%S") if field.retrieved_at else "",
                field.note or "",
            ])
            if field.disputed:
                for c in range(1, len(prov_headers) + 1):
                    ws2.cell(row=ws2.max_row, column=c).fill = _DISPUTE_FILL
    _autosize(ws2, max_width=60)

    # --- Sheet 3: discrepancies --------------------------------------------
    ws3 = wb.create_sheet("Discrepancies")
    disc_headers = ["VIN", "Field", "Severity", "Selected Value", "Selected Source",
                    "Conflicting Value", "Conflicting Source", "Conflicting Confidence"]
    _style_header(ws3, disc_headers)
    any_discrepancy = False
    for record in records:
        for d in record.discrepancies:
            for conflict in d.conflicting:
                any_discrepancy = True
                ws3.append([
                    record.vin, d.label, d.severity, _cell(d.selected_value),
                    d.selected_source, _cell(conflict.value), conflict.source,
                    conflict.confidence.value,
                ])
    if not any_discrepancy:
        ws3.append(["No discrepancies detected across sources for this export."])
    _autosize(ws3, max_width=60)

    # --- Sheet 4: provenance summary ---------------------------------------
    ws4 = wb.create_sheet("Export Info")
    ws4["A1"] = "VIN Decoder export"
    ws4["A1"].font = _TITLE_FONT
    meta = [
        ("Generated (UTC)", _dt.datetime.now(_dt.UTC).strftime("%Y-%m-%d %H:%M:%S")),
        ("Vehicles", len(records)),
        ("Valid", sum(1 for r in records if r.valid)),
        ("With discrepancies", sum(1 for r in records if r.discrepancies)),
        ("Served from cache", sum(1 for r in records if r.cached)),
        ("Sources used", ", ".join(sorted({s for r in records for s in r.sources})) or "None"),
        ("", ""),
        ("Note", "Highlighted rows had conflicting source data (amber) or failed "
                 "validation (red). Fields marked 'Not available' were not supplied "
                 "by any source and have not been estimated."),
    ]
    for i, (key, value) in enumerate(meta, start=3):
        ws4.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws4.cell(row=i, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_filename(extension: str, prefix: str = "vin-decode") -> str:
    stamp = _dt.datetime.now(_dt.UTC).strftime("%Y%m%d-%H%M%S")
    return f"{prefix}-{stamp}.{extension}"
